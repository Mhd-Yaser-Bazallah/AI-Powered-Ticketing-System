from pathlib import Path
from typing import Optional


def extract_text(file_path: str, kind: Optional[str] = None) -> str:
    suffix = Path(file_path).suffix.lower()
    if kind == "pdf" or suffix == ".pdf":
        return _extract_pdf(file_path)
    if kind == "word" or suffix in {".doc", ".docx"}:
        return _extract_docx(file_path)
    if kind == "excel" or suffix in {".xls", ".xlsx"}:
        return _extract_xlsx(file_path)
    raise ValueError(f"Unsupported file type: kind={kind} suffix={suffix}")


def _extract_pdf(file_path: str) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("Missing dependency: pypdf") from exc

    reader = PdfReader(file_path)
    parts = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return "\n".join(parts)


def _extract_docx(file_path: str) -> str:
    try:
        from docx import Document
    except Exception as exc:
        raise RuntimeError("Missing dependency: python-docx") from exc

    doc = Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(file_path: str) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl") from exc

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            line = " | ".join(v for v in values if v != "")
            if line:
                parts.append(line)
    return "\n".join(parts)
