                 
                                      
                               

                                                             
                                    
                             

                   
                                                   
                                                         

                
                                                  
                                                            
                                                                       

                        
                              
                                                                                          
       
                                                                          

                             
                     



import openpyxl
import os
from django.http import JsonResponse
from django.conf import settings

def export_to_excel(headers, rows, filename="report.xlsx"):

                               
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

                   
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

                     
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

                             
    report_folder = os.path.join(settings.MEDIA_ROOT, "reports")
    os.makedirs(report_folder, exist_ok=True)

                          
    file_path = os.path.join(report_folder, filename)
    workbook.save(file_path)

                                            
    file_url = f"{settings.MEDIA_URL}reports/{filename}"

    return JsonResponse({"download_url": file_url}, status=200)